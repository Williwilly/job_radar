#!/usr/bin/env python3
"""
job_search_automation_excel_status.py
====================================

Multi-source job-search automation (RSS -> Excel .xlsx) with a STATUS dropdown
in the "Seen Jobs" sheet that preserves your manual edits.

What it does
------------
- Pulls job postings from multiple RSS feeds
- Filters by keywords
- Deduplicates across feeds and across runs
- Writes to jobs.xlsx with:
    * "New Jobs"  -> only jobs found in the current run
    * "Seen Jobs" -> running history of all captured jobs
- Adds a Status dropdown in Seen Jobs:
    Not Applied / Applied / Interview / Rejected
- Does NOT overwrite your existing status selections or other manual edits

Run
---
python job_search_automation_excel_status.py

Optional examples
-----------------
python job_search_automation_excel_status.py --keywords "product analyst,data analyst,business intelligence"
python job_search_automation_excel_status.py --exclude-keywords "senior,staff,principal,lead,manager,head,director,vp"
python job_search_automation_excel_status.py --xlsx-output jobs.xlsx
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import ssl
import sys
import time
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from typing import Dict, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# -----------------------------
# Defaults
# -----------------------------

DEFAULT_FEEDS: List[str] = [
    "https://remotive.com/remote-jobs/feed/data",
    "https://remotive.com/remote-jobs/feed/product",
    "https://weworkremotely.com/categories/remote-product-jobs.rss",
    "https://remoteok.com/remote-jobs.rss",
]

DEFAULT_KEYWORDS: List[str] = [
    "product analyst",
    "data analyst",
    "business intelligence",
    "business intelligence analyst",
    "bi analyst",
    "bi developer",
    "product analytics",
    "analytics engineer",
    "insights analyst",
    "reporting analyst",
]

DEFAULT_EXCLUDE_KEYWORDS: List[str] = [
    "vp",
    "vice president",
    "director",
]

STATUS_OPTIONS: List[str] = [
    "Not Applied",
    "Applied",
    "Interview",
    "Rejected",
]

SHEET_NEW = "New Jobs"
SHEET_SEEN = "Seen Jobs"

BASE_COLUMNS: List[str] = [
    "run_ts_utc",
    "source",
    "title",
    "link",
    "published",
    "matched_keywords",
    "description",
    "guid",
]


# -----------------------------
# Model
# -----------------------------

@dataclass(frozen=True)
class JobItem:
    run_ts_utc: str
    source: str
    title: str
    link: str
    published: str
    matched_keywords: str
    description: str
    guid: str


# -----------------------------
# Basic helpers
# -----------------------------

def _now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_whitespace(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def strip_html(s: str) -> str:
    if not s:
        return ""
    s = re.sub(r"(?is)<script.*?>.*?</script>", " ", s)
    s = re.sub(r"(?is)<style.*?>.*?</style>", " ", s)
    s = re.sub(r"(?s)<.*?>", " ", s)
    return normalize_whitespace(s)


def parse_pubdate_rss(item: ET.Element) -> str:
    pub = (item.findtext("pubDate") or "").strip()
    if not pub:
        return ""
    try:
        dt = parsedate_to_datetime(pub)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).isoformat(timespec="seconds")
    except Exception:
        return ""


def build_stable_id(source: str, guid: str, link: str, title: str) -> str:
    base = "|".join([source, guid or "", link or "", title or ""])
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def parse_csv_list(value: Optional[str]) -> List[str]:
    if not value:
        return []
    return [p.strip() for p in value.split(",") if p.strip()]


def compile_keywords(keywords: Sequence[str]) -> List[str]:
    out = []
    for k in keywords:
        k2 = normalize_whitespace(k).lower()
        if k2:
            out.append(k2)
    return sorted(set(out))


def matches_keywords(title: str, description: str, include: Sequence[str], exclude: Sequence[str]) -> Tuple[bool, List[str]]:
    blob = f"{title}\n{description}".lower()
    matched = [k for k in include if k in blob]
    if not matched:
        return False, []
    for x in exclude:
        if x and x in blob:
            return False, []
    return True, matched


def load_seen(path: str) -> Set[str]:
    if not os.path.exists(path):
        return set()
    with open(path, "r", encoding="utf-8") as f:
        return {line.strip() for line in f if line.strip()}


def save_seen(path: str, ids: Set[str]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for _id in sorted(ids):
            f.write(_id + "\n")


# -----------------------------
# Fetch + parse RSS
# -----------------------------

HEADER_PROFILES: List[Dict[str, str]] = [
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Accept": "application/rss+xml, application/xml;q=0.9, text/xml;q=0.8, */*;q=0.1",
        "Accept-Language": "en-US,en;q=0.9",
    },
    {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 "
                      "(KHTML, like Gecko) Version/17.0 Safari/605.1.15",
        "Accept": "application/rss+xml, application/xml;q=0.9, text/xml;q=0.8, */*;q=0.1",
        "Accept-Language": "en-US,en;q=0.9",
    },
    {"User-Agent": "curl/8.0.1", "Accept": "*/*"},
]


def fetch_url(url: str, timeout: int = 30, retries: int = 3, backoff_sec: float = 1.5) -> str:
    last_err: Optional[Exception] = None
    ctx = ssl.create_default_context()

    for attempt in range(1, retries + 1):
        for headers in HEADER_PROFILES:
            req = urllib.request.Request(url, headers=headers)
            try:
                with urllib.request.urlopen(req, context=ctx, timeout=timeout) as resp:
                    raw = resp.read()
                try:
                    return raw.decode("utf-8")
                except UnicodeDecodeError:
                    return raw.decode("latin-1", errors="replace")
            except Exception as e:
                last_err = e
                continue
        time.sleep(backoff_sec * attempt)

    assert last_err is not None
    raise last_err


def parse_rss(xml_text: str) -> List[Tuple[str, str, str, str, str]]:
    root = ET.fromstring(xml_text)
    channel = root.find("channel")
    if channel is None:
        return []
    out: List[Tuple[str, str, str, str, str]] = []
    for it in channel.findall("item"):
        title = normalize_whitespace(it.findtext("title") or "")
        link = normalize_whitespace(it.findtext("link") or "")
        guid = normalize_whitespace(it.findtext("guid") or "")
        pub = parse_pubdate_rss(it)
        desc = strip_html(it.findtext("description") or it.findtext("content:encoded") or "")
        out.append((title, link, pub, desc, guid))
    return out


# -----------------------------
# Excel helpers
# -----------------------------

def _find_header_col(ws, header_name: str) -> Optional[int]:
    """
    Find a column by header text in row 1.
    """
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if isinstance(value, str) and value.strip().lower() == header_name.lower():
            return col_idx
    return None


def _ensure_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    if SHEET_NEW not in wb.sheetnames:
        wb.create_sheet(SHEET_NEW)
    if SHEET_SEEN not in wb.sheetnames:
        wb.create_sheet(SHEET_SEEN)

    return wb


def _ensure_new_headers(ws) -> None:
    """
    Force New Jobs to use the base schema only.
    """
    for i, name in enumerate(BASE_COLUMNS, start=1):
        ws.cell(row=1, column=i, value=name)


def _ensure_seen_headers(ws) -> int:
    """
    Ensure base columns exist in columns 1..8.
    Ensure a 'status' header exists somewhere and return that column index.
    We do NOT overwrite any existing status values.
    """
    for i, name in enumerate(BASE_COLUMNS, start=1):
        ws.cell(row=1, column=i, value=name)

    status_col = _find_header_col(ws, "status")
    if status_col is None:
        # Prefer column right after base columns if empty, otherwise append at end
        preferred = len(BASE_COLUMNS) + 1
        existing = ws.cell(row=1, column=preferred).value
        if existing in (None, ""):
            status_col = preferred
        else:
            status_col = ws.max_column + 1
        ws.cell(row=1, column=status_col, value="status")

    return status_col


def _sheet_has_data(ws) -> bool:
    return ws.max_row >= 2 and any(
        ws.cell(row=2, column=c).value is not None for c in range(1, ws.max_column + 1)
    )


def _clear_sheet_except_header(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _read_existing_guids(ws) -> Set[str]:
    guid_col = _find_header_col(ws, "guid") or len(BASE_COLUMNS)
    out: Set[str] = set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=guid_col).value
        if isinstance(val, str) and val.strip():
            out.add(val.strip())
    return out


def _append_jobs_new(ws, jobs: List[JobItem]) -> None:
    """
    Append rows to New Jobs sheet.
    """
    link_col = BASE_COLUMNS.index("link") + 1
    for j in jobs:
        ws.append([j.run_ts_utc, j.source, j.title, j.link, j.published, j.matched_keywords, j.description, j.guid])
        r = ws.max_row
        cell = ws.cell(row=r, column=link_col)
        if j.link:
            cell.hyperlink = j.link
            cell.style = "Hyperlink"


def _append_jobs_seen(ws, jobs: List[JobItem], status_col: int) -> None:
    """
    Append rows to Seen Jobs without touching existing rows or existing status cells.
    New rows get default status = 'Not Applied'.
    """
    link_col = BASE_COLUMNS.index("link") + 1
    for j in jobs:
        row_num = ws.max_row + 1
        values = [j.run_ts_utc, j.source, j.title, j.link, j.published, j.matched_keywords, j.description, j.guid]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        if j.link:
            cell = ws.cell(row=row_num, column=link_col)
            cell.hyperlink = j.link
            cell.style = "Hyperlink"
        # Only set status for the new row
        ws.cell(row=row_num, column=status_col, value=STATUS_OPTIONS[0])


def _apply_status_dropdown(ws, status_col: int) -> None:
    """
    Apply/refresh the dropdown on the entire status column from row 2 downward.
    Existing values remain unchanged.
    """
    formula = '"' + ",".join(STATUS_OPTIONS) + '"'

    # Remove old copies of our exact dropdown (to avoid stacking duplicates)
    existing = list(ws.data_validations.dataValidation)
    ws.data_validations.dataValidation = [
        dv for dv in existing
        if not (dv.type == "list" and dv.formula1 == formula)
    ]

    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.prompt = "Select application status"
    dv.error = "Choose a valid status from the dropdown."
    ws.add_data_validation(dv)

    col_letter = get_column_letter(status_col)
    dv.add(f"{col_letter}2:{col_letter}1048576")


def _apply_pretty_format_new(ws) -> None:
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(BASE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(BASE_COLUMNS))}1"

    widths = {
        "run_ts_utc": 20,
        "source": 35,
        "title": 45,
        "link": 45,
        "published": 20,
        "matched_keywords": 28,
        "description": 70,
        "guid": 14,
    }
    for col_idx, col_name in enumerate(BASE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)

    desc_col = BASE_COLUMNS.index("description") + 1
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=desc_col).alignment = Alignment(vertical="top", wrap_text=True)


def _apply_pretty_format_seen(ws, status_col: int) -> None:
    header_font = Font(bold=True)

    # Base headers
    for col_idx, col_name in enumerate(BASE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Status header
    status_header = ws.cell(row=1, column=status_col)
    status_header.value = "status"
    status_header.font = header_font
    status_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    last_col = max(status_col, len(BASE_COLUMNS))
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}1"

    widths = {
        "run_ts_utc": 20,
        "source": 35,
        "title": 45,
        "link": 45,
        "published": 20,
        "matched_keywords": 28,
        "description": 70,
        "guid": 14,
        "status": 16,
    }
    for col_idx, col_name in enumerate(BASE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)
    ws.column_dimensions[get_column_letter(status_col)].width = widths["status"]

    desc_col = BASE_COLUMNS.index("description") + 1
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=desc_col).alignment = Alignment(vertical="top", wrap_text=True)


def seen_sheet_is_empty(xlsx_path: str) -> bool:
    if not os.path.exists(xlsx_path):
        return True
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        if SHEET_SEEN not in wb.sheetnames:
            return True
        ws = wb[SHEET_SEEN]
        return not _sheet_has_data(ws)
    except Exception:
        return True


def write_excel(xlsx_path: str, new_jobs: List[JobItem], history_add_jobs: List[JobItem]) -> None:
    wb = _ensure_workbook(xlsx_path)
    ws_new = wb[SHEET_NEW]
    ws_seen = wb[SHEET_SEEN]

    _ensure_new_headers(ws_new)
    status_col = _ensure_seen_headers(ws_seen)

    # New Jobs = this run only
    _clear_sheet_except_header(ws_new)
    _append_jobs_new(ws_new, new_jobs)

    # Seen Jobs = cumulative history, append only truly new-to-workbook rows
    existing_guids = _read_existing_guids(ws_seen)
    to_add = [j for j in history_add_jobs if j.guid not in existing_guids]
    _append_jobs_seen(ws_seen, to_add, status_col=status_col)

    _apply_status_dropdown(ws_seen, status_col=status_col)
    _apply_pretty_format_new(ws_new)
    _apply_pretty_format_seen(ws_seen, status_col=status_col)

    wb.save(xlsx_path)


# -----------------------------
# Main
# -----------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Multi-feed RSS job search automation (Excel + status dropdown).")
    parser.add_argument("--feeds", default=",".join(DEFAULT_FEEDS), help="Comma-separated RSS feed URLs.")
    parser.add_argument("--keywords", default=",".join(DEFAULT_KEYWORDS), help="Comma-separated include keywords.")
    parser.add_argument("--exclude-keywords", default=",".join(DEFAULT_EXCLUDE_KEYWORDS), help="Comma-separated exclude keywords.")
    parser.add_argument("--xlsx-output", default="jobs.xlsx", help="Excel output path.")
    parser.add_argument("--seen-file", default="seen.txt", help="Seen IDs file.")
    parser.add_argument("--max-new", type=int, default=200, help="Safety cap: max NEW jobs per run.")
    parser.add_argument("--debug", action="store_true", help="Print feed errors.")
    args = parser.parse_args()

    feeds = [f.strip() for f in args.feeds.split(",") if f.strip()]
    include_kw = compile_keywords(parse_csv_list(args.keywords))
    exclude_kw = compile_keywords(parse_csv_list(args.exclude_keywords))

    if not feeds:
        raise SystemExit("No feeds provided.")
    if not include_kw:
        raise SystemExit("No keywords provided.")

    seen = load_seen(args.seen_file)
    run_ts = _now_utc_iso()

    all_candidates: List[JobItem] = []
    errors: List[str] = []

    for feed_url in feeds:
        try:
            xml_text = fetch_url(feed_url)
            items = parse_rss(xml_text)
            for (title, link, published, desc, guid_raw) in items:
                if not title and not link:
                    continue
                ok, matched = matches_keywords(title, desc, include=include_kw, exclude=exclude_kw)
                if not ok:
                    continue
                stable_id = build_stable_id(feed_url, guid_raw, link, title)
                all_candidates.append(
                    JobItem(
                        run_ts_utc=run_ts,
                        source=feed_url,
                        title=title,
                        link=link,
                        published=published,
                        matched_keywords=",".join(matched),
                        description=desc,
                        guid=stable_id,
                    )
                )
        except Exception as e:
            msg = f"Feed failed: {feed_url} -> {type(e).__name__}: {e}"
            errors.append(msg)
            if args.debug:
                print(msg, file=sys.stderr)

    # Unique by guid
    uniq_by_id: Dict[str, JobItem] = {}
    for j in all_candidates:
        uniq_by_id.setdefault(j.guid, j)

    candidates = list(uniq_by_id.values())
    new_jobs = [j for j in candidates if j.guid not in seen]

    # Sort newest first where possible
    def sort_key(job: JobItem) -> Tuple[int, str]:
        return (0, job.published) if job.published else (1, "")
    candidates.sort(key=sort_key, reverse=True)
    new_jobs.sort(key=sort_key, reverse=True)

    if len(new_jobs) > args.max_new:
        new_jobs = new_jobs[: args.max_new]

    # If Seen Jobs sheet is empty, seed it with current matches so workbook is never blank
    seed_needed = seen_sheet_is_empty(args.xlsx_output)
    history_add = candidates if seed_needed else new_jobs

    write_excel(args.xlsx_output, new_jobs=new_jobs, history_add_jobs=history_add)

    # Keep seen.txt consistent with what exists in Seen Jobs history
    seen.update({j.guid for j in history_add})
    save_seen(args.seen_file, seen)

    print(f"Feeds checked: {len(feeds)}")
    print(f"Matched candidates (pre-dedupe): {len(all_candidates)}")
    print(f"Unique matched: {len(uniq_by_id)}")
    print(f"NEW written this run: {len(new_jobs)}")
    print(f"Excel: {args.xlsx_output}")
    print(f"Seen file: {args.seen_file}")
    if errors and not args.debug:
        print(f"Feeds failed: {len(errors)} (run with --debug to see details)", file=sys.stderr)


if __name__ == "__main__":
    main()
